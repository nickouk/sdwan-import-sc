#!/usr/bin/env python3


# This code will be customer specific as each customer will have different template variables
#
# import tracker sheet and build template csv for import into vManage to cutdown on manual work required to deploy routers
# To adapt this code there are two main sections that require updating:
# Section 1 is the definition of vmanage_dict - each dictionary key maps to a column header which is a variable in a template
# Section 2 is the main loop that reads the tracker sheet, manipulates the data and then writes it into the dictionary
# Section 3 performs postcode lookups to obtain GPS coords and gathers a list of routes required for DNAC - the vmanage-import-[cust].csv file is written


# openpxyl is a library for handing MS Excel files

import openpyxl

# pandas is used for working with csv files

import pandas as pd

# requests allows API calls - used to correct the UK Postcodes which have no space

import requests

# allows whois lookup for public subnets ; static routes are required on DNAC

from ipwhois import IPWhois

# some standard libraries

import pprint
import ipaddress
import sys
import re
import math
import pickle
import json
import math

def store_nets(store_num):

	# generate store networks from store number

    # python splice uses start:end so to get characters 1-3 use 0:3 negative indexes work from end of string  
    # 1:4 gets characters 2nd, 3rd and 4th characters
    actual_store_num = int(store_num[1:4])  # strip leading digit for network calc
    
    # 0:2 gets first two characters
    store_net_oct2 = store_num[0:2]
    if store_net_oct2[0] == '0' or store_net_oct2[0] == '9':
        # 1: starts at character 2 to end of string
        store_net_oct2 = store_net_oct2[1:]

    store_net_oct2 = int(store_net_oct2)
    store_net_oct3 = int(store_num[2:4])

    store_net_oct2_vlan70 = store_net_oct2
    store_net_oct2_vlan31 = (f'{store_net_oct2:1>3}')
    store_net_oct3_vlan31 = store_net_oct3
    store_net_oct2_vlan101 = (f'{store_net_oct2:1>3}')

    if actual_store_num < 255:
        store_net_oct2 = 1
        store_net_oct2_vlan70 = 100
        store_net_oct3 = actual_store_num

    return(store_net_oct2, store_net_oct3, store_net_oct2_vlan70, store_net_oct2_vlan31, store_net_oct3_vlan31, store_net_oct2_vlan101)

def test_store_nets():

	# --- Test code to generate SCOOP store subnets ---
	keys = ['store','vlan60','vlan70','vlan10','vlan20','vlan31','vlan101']
	subnets_dict = {key: [] for key in keys}
	
	file1 = open('/mnt/c/Users/nick.oneill/Downloads/subnets.txt', 'w')
	# test generated store numbers from 0000 to 9999
	
	for store in range(0, 9999):
	
		store_num = str(store).zfill(4)
		store_net_oct2, store_net_oct3, store_net_oct2_vlan70, store_net_oct2_vlan31, store_net_oct3_vlan31, store_net_oct2_vlan101 = store_nets(store_num)
		
		# print(f'store_num {store_num} store_net_oct2 {store_net_oct2} store_net_oct3 {store_net_oct3}')
		try:
			vlan60_ipv4 = ipaddress.ip_network(f'151.{store_net_oct2}.{store_net_oct3}.0/24')
		except:
			vlan60_ipv4 = (f'{store} error')
		try:
			vlan70_ipv4 = ipaddress.ip_network(f'10.{store_net_oct2_vlan70}.{store_net_oct3}.0/24')
		except:
			vlan70_ipv4 = (f'{store} error')
		try:
			vlan10_ipv4 = ipaddress.ip_network(f'10.1{store_net_oct2}.{store_net_oct3}.0/25')
		except:
			vlan10_ipv4 = (f'{store} error')
		try:
			vlan20_ipv4 = ipaddress.ip_network(f'10.1{store_net_oct2}.{store_net_oct3}.128/25')
		except:
			vlan20_ipv4 = (f'{store} error')
		try:
			vlan31_ipv4 = ipaddress.ip_network(f'10.{store_net_oct2_vlan31}.{store_net_oct3_vlan31}.0/28')
		except:
			vlan31_ipv4 = (f'{store} error')
		try:
			vlan101_ipv4 = ipaddress.ip_network(f'10.{store_net_oct2_vlan101}.{store_net_oct3}.224/27')
		except:
			vlan101_ipv4 = (f'{store} error')
	
		subnets_dict['store'].append(f'{store:0>4}')
		subnets_dict['vlan60'].append(str(vlan60_ipv4))
		subnets_dict['vlan70'].append(str(vlan70_ipv4))
		subnets_dict['vlan10'].append(str(vlan10_ipv4))
		subnets_dict['vlan20'].append(str(vlan20_ipv4))
		subnets_dict['vlan31'].append(str(vlan31_ipv4))
		subnets_dict['vlan101'].append(str(vlan101_ipv4))
	
	df = pd.DataFrame(subnets_dict)
	
	try:
		df.to_csv('/mnt/c/Users/nick.oneill/Downloads/scoop-subnets.csv', index=False)
	except PermissionError:
		print('*' * 120,'\nError: scoop-subnets.csv is open in another application - please close and re-run the script\n','*' * 120)
		exit()
	
	# all done
	print('scoop-subnets.csv has been created :)\n')
	file1.close()
	return()

def postcode_api(postcode_apilist):

    # Function for passing a list of postcodes to an external site for lookup
    # Correct the postcode format (missing space) and return long + lat values
    # Send API request, passing in postcodes as a list
    # NOTE Maximum 100 postcodes

    postcode_uri = 'https://api.postcodes.io/postcodes'

    # Raise an exception requests.HTTPException error is response is anything other than 200 (OK)

    json = {"postcodes": postcode_apilist}
    postcode_lookup = ''
    try:
        postcode_lookup = requests.post(
            postcode_uri,
            json={"postcodes": postcode_apilist}
        )
        postcode_lookup.raise_for_status()
    except requests.exceptions.ConnectionError:
        print(f'\nConnection error connecting to {postcode_uri}\nvManage import sheet has not been updated\n')
        sys.exit()
    except requests.HTTPError as error:
        print(f'\nHTTP Error:\n{error}')
        sys.exit()

    return (postcode_lookup)


def circuit_bandwidth(circuit_type):

    # Function to return circuit bandwidth based on circuit type

    if circuit_type == 'FTTP':
        return (80, 20, 'GigabitEthernet0/0/0')
    elif circuit_type == 'SOGEA':
        return (80, 20, 'Dialer1')
    elif circuit_type == 'FTTC':
        return (80, 20, 'Dialer1')
    elif circuit_type == 'ADSL':
        return (24, 3, 'Dialer1')
    elif circuit_type == 'OFNL Fibre':
        return (80, 20, 'GigabitEthernet0/0/0')
    else:
        # return a tuple so callers that unpack won't fail
        return (0, 0, 0)


def sanatise_serial(serial):

    # Function to check the location code is 3 letters and remove leading 'S'
    if not serial:
        return ''
    # guard against short serials
    if len(serial) > 3:
        serial_check = serial[3]  # check 4th character is a number
    else:
        serial_check = ''
    if serial_check.isalpha() and serial[0].upper() == 'S':
        return (serial[1:].upper())
    else:
        return (serial.upper())

def wan_color(circuit_provider):

    # Function to return wan color based on circuit provider
    
    if circuit_provider == 'BT' or circuit_provider == 'MAINTEL-BT':
        return 'blue'
    elif circuit_provider == 'PXC' or circuit_provider == 'MAINTEL-PXC':
        return 'green'
    elif circuit_provider == 'Other':
        return 'public-internet'
    else:
        return None

# -----------------------------
# --- Main code starts here ---
# -----------------------------
# Open the tracker sheet

try:
    tracker_wb_obj = openpyxl.load_workbook(
        '/mnt/c/Users/nick.oneill/Downloads/NOF2025 Rollout tracker.xlsx')
except FileNotFoundError:
    print('*' * 120,'\nError: NOF2025 Rollout tracker.xlsx file not found - please download from Sharepoint and re-run the script\n','*' * 120)
    sys.exit()

tracker_sheet_obj = tracker_wb_obj.active
# determine how many rows we have
max_row = tracker_sheet_obj.max_row

# initialise some variables
keys = ['Device ID',
'System IP',
'Host Name',
'Site Id',
'Dual Stack IPv6 Default',
'Rollback Timer (sec)',
'basic_gpsl_longitude',
'basic_gpsl_latitude',
'provision_port_disable',
'vlan31_vrrp_pri',
'vlan31_vrrp_ipv4',
'vlan31_ipv4',
'vlan31_mask',
'vlan31_dhcp_net',
'vlan31_dhcp_mask',
'vlan31_dhcp_exclude',
'vlan31_dhcp_gateway',
'vlan120_vrrp_pri',
'vlan120_vrrp_ipv4',
'vlan120_ipv4',
'vlan120_mask',
'vlan120_dhcp_exclude',
'vlan100_vrrp_pri',
'vlan100_vrrp_ipv4',
'vlan100_ipv4',
'vlan100_mask',
'vlan100_dhcp_exclude',
'vlan101_vrrp_pri',
'vlan101_vrrp_ipv4',
'vlan101_ipv4',
'vlan101_mask',
'vlan101_dhcp_net',
'vlan101_dhcp_mask',
'vlan101_dhcp_exclude',
'vlan101_dhcp_gateway',
'vlan40_vrrp_pri',
'vlan40_vrrp_ipv4',
'vlan40_ipv4',
'vlan40_mask',
'vlan40_dhcp_exclude',
'vlan30_vrrp_pri',
'vlan30_vrrp_ipv4',
'vlan30_ipv4',
'vlan30_mask',
'vlan30_dhcp_exclude',
'lan_vpn_100_nat_1_rangeStart',
'lan_vpn_100_nat_1_rangeEnd',
'lan_vpn_100_staticNat_1_translatedSourceIp',
'lan_vpn_100_staticNat_2_translatedSourceIp',
'loopback0_ipv4',
'loopback0_mask',
'vlan20_vrrp_pri',
'vlan20_vrrp_ipv4',
'vlan20_ipv4',
'vlan20_mask',
'vlan20_dhcp_net',
'vlan20_dhcp_mask',
'vlan20_dhcp_exclude',
'vlan20_dhcp_gateway',
'vlan10_vrrp_pri',
'vlan10_vrrp_ipv4',
'vlan10_ipv4',
'vlan10_mask',
'vlan10_dhcp_net',
'vlan10_dhcp_mask',
'vlan10_dhcp_exclude',
'vlan10_dhcp_gateway',
'vlan2_vrrp_pri',
'vlan2_vrrp_ipv4',
'vlan2_ipv4',
'vlan2_mask',
'vlan2_dhcp_net',
'vlan2_dhcp_mask',
'vlan2_dhcp_exclude',
'vlan2_dhcp_gateway',
'vlan80_vrrp_pri',
'vlan80_vrrp_ipv4',
'vlan80_ipv4',
'vlan80_mask',
'vlan70_vrrp_pri',
'vlan70_vrrp_ipv4',
'vlan70_ipv4',
'vlan70_mask',
'vlan60_vrrp_pri',
'vlan60_vrrp_ipv4',
'vlan60_ipv4',
'vlan60_mask',
'tloc_next_hop',
'tloc_bandwidth_up',
'tloc_bandwidth_down',
'wan_bandwidth_up',
'wan_bandwidth_down',
'wan_desc',
'ethpppoe_chapHost',
'ethpppoe_chapPwd',
'wan_color',
'ethpppoe_ipsecPrefer',
'wan_shapingRate',
'wan_track_addr',
'wan_track_addr_tloc',
'cloudSaaSDeviceRole_variable',
'cloudSaaSVpnType_variable',
'cloudSaasSigTunnelList_variable',
'cloudSaasTlocList_variable',
'cloudSaasSigEnabled_variable',
'cloudSaasInterfaceList_variable',
'cloudSaasLBEnabled_variable',
'cloudSaasLoss_variable',
'cloudSaasLatency_variable',
'cloudSaasSourceIpBased_variable',
'qos_Interface_1']

vmanage_dict = {key: [] for key in keys}

# define column numbers for the tracker sheet (1 = column A) - this makes it easier to modify later if the tracker sheet changes
store_num_col = 1  # column A
store_type_col = 2  # column B
postcode_col = 4  # column D
router1_serial_col = 5  # column E
router1_mgmt_ip_col = 6  # column F
circuit1_provider_col =  7  # column G
circuit1_type_col = 8  # column H
circuit1_bw_up_col = 9  # column I
circuit1_bw_down_col = 10  # column J
circuit1_ref_col = 11  # column K
circuit1_ppp_name_col = 13  # column M
circuit1_ppp_pwd_col = 14  # column N
router2_serial_col = 15  # column O
router2_mgmt_ip_col = 16  # column P
circuit2_provider_col = 17  # column Q
circuit2_type_col = 18  # column R
circuit2_bw_up_col = 19  # column S
circuit2_bw_down_col = 20  # column T
circuit2_ref_col = 21  # column U
circuit2_ppp_name_col = 23  # column W
circuit2_ppp_pwd_col = 24  # column X
vlan2_col = 25  # column Y
vlan60_col = 26  # column Z
provision_port_disable_col = 27 # column AA

unique_subnets = set()

# main loop - loop through the tracker sheet and build rows for the vmanage-import-sc.csv dictionary transforming some of the data

test_run = False  # set to True to test store_nets function only

if test_run:
    print('\nTest run selected - no changes will be made to vManage import sheet\n')
    novalue = test_store_nets()
    sys.exit()

tracker_row = 3
postcode_list = []
print(f'{max_row} rows found ...\n')

while tracker_row <= max_row:

    # get the store number and pad to 4 digits
    store_num = str(tracker_sheet_obj.cell(row=tracker_row, column=store_num_col).value).zfill(4)

    # if store number is missing skip to next row
    if store_num == '0000' or store_num == 'None':
        tracker_row = tracker_row + 1
        continue

    # get the store type
    store_type = str(tracker_sheet_obj.cell(row=tracker_row, column=store_type_col).value).upper()
    store_type = int(store_type[0])  # first character only
    site_id = f'{store_type}{store_num}'

    # get the postcode
    postcode = str(tracker_sheet_obj.cell(row=tracker_row, column=postcode_col).value).upper().replace(' ', '')
    # moved the append postcode to after router checks to avoid arrys being different sizes


    # get router 1 serial number
    router1_serial = str(tracker_sheet_obj.cell(row=tracker_row, column=router1_serial_col).value).upper()
    router1_serial = sanatise_serial(router1_serial)

    if router1_serial == 'NONE' or router1_serial == '':
        print(f'Error: missing router 1 serial number for store {store_num} row {tracker_row}  ... skipping to next row')
        tracker_row = tracker_row + 1
        continue
  
    # get circuit 1 type and bandwidth
    circuit1_type = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit1_type_col).value).upper()
    circuit1_bw_down = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit1_bw_down_col).value)
    circuit1_bw_up = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit1_bw_up_col).value)
    
    if circuit1_bw_down == 'None' or circuit1_bw_up == 'None':
        circuit1_bw_down, circuit1_bw_up, interface1 = circuit_bandwidth(circuit1_type)
    else:
        a, b, interface1 = circuit_bandwidth(circuit1_type)
    
    circuit1_bw_down = float(circuit1_bw_down)
    circuit1_bw_up = float(circuit1_bw_up)

    circuit1_ref = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit1_ref_col).value).upper()

    # initialize router 2 variables in case of a single site router
    router2_serial = 'NONE'
    circuit2_type = 'NONE'
    circuit2_bw_down = 0
    circuit2_bw_up = 0
    circuit2_type = 'NONE'
    circuit2_ref = 'NONE'
    circuit2_ppp_name = 'NONE'
    circuit2_ppp_pwd = 'NONE'
    router2_mgmt_ip = 'NONE'
    router2_systemip = 'NONE'
    router2_hostname = 'NONE'
    router2_wan_color = 'NONE'

    # get router 2 serial number if present otherwsie assume a singe router site
    router2_serial = str(tracker_sheet_obj.cell(row=tracker_row, column=router2_serial_col).value).upper()
    if router2_serial != 'NONE':
        router2_serial = sanatise_serial(router2_serial)

        # get circuit 2 type and bandwidth
        circuit2_type = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit2_type_col).value).upper()
        circuit2_bw_down = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit2_bw_down_col).value)
        circuit2_bw_up = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit2_bw_up_col).value)

        if circuit2_bw_down == 'None' or circuit2_bw_up == 'None':
            circuit2_bw_down, circuit2_bw_up, interface2 = circuit_bandwidth(circuit2_type)
        else:
            a, b, interface2 = circuit_bandwidth(circuit2_type)
        
        circuit2_bw_down = float(circuit2_bw_down)
        circuit2_bw_up = float(circuit2_bw_up)

        circuit2_ref = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit2_ref_col).value).upper()

        # get managment IP address for router 2
        router2_mgmt_ip = str(tracker_sheet_obj.cell(row=tracker_row, column=router2_mgmt_ip_col).value)
        if '/' not in router2_mgmt_ip: router2_mgmt_ip = router2_mgmt_ip + '/32'
        router2_mgmt_ip = ipaddress.ip_network(router2_mgmt_ip, strict=False)
        router2_systemip = router2_mgmt_ip.network_address

        # build router 2 hostname
        router2_hostname = f'SC-{store_type}-{store_num}-R2'

        # get provider for circuit 2
        circuit2_provider = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit2_provider_col).value).upper()
        router2_wan_color = wan_color(circuit2_provider)

        # get ppp name and password for circuit 2
        circuit2_ppp_name = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit2_ppp_name_col).value)
        circuit2_ppp_pwd = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit2_ppp_pwd_col).value)

        if circuit2_provider == 'BT' and circuit2_ppp_name == 'None':
            circuit2_ppp_name = 'dummy@bband1.com'
            circuit2_ppp_pwd = 'dummy'

        if circuit2_ppp_name == 'None':
            circuit2_ppp_name = 'notrequired'

        # we need to duplicate the postcode for router 2 which is not optimal but works for now
        postcode_list.append(postcode)

    # get managment IP address for router 1
    router1_mgmt_ip = str(tracker_sheet_obj.cell(row=tracker_row, column=router1_mgmt_ip_col).value)
    if '/' not in router1_mgmt_ip: router1_mgmt_ip = router1_mgmt_ip + '/32'
    try:
        router1_mgmt_ip = ipaddress.ip_network(router1_mgmt_ip, strict=False)
    except ValueError:
        print(f'Error: invalid management IP address for store {store_num} row {tracker_row}  ... skipping to next row')
        tracker_row = tracker_row + 1
        continue

    router1_systemip = router1_mgmt_ip.network_address

    # only safe to append post code if we are not skipping the row (we fetched the post code earlier in the loop)
    postcode_list.append(postcode)

    # build router 1 hostname
    router1_hostname = f'SC-{store_type}-{store_num}-R1'

    # get provider for circuit 1
    circuit1_provider = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit1_provider_col).value).upper()
    router1_wan_color = wan_color(circuit1_provider)

    # get ppp name and password for circuit 1
    circuit1_ppp_name = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit1_ppp_name_col).value)
    circuit1_ppp_pwd = str(tracker_sheet_obj.cell(row=tracker_row, column=circuit1_ppp_pwd_col).value)

    if circuit1_provider == 'BT' and circuit1_ppp_name == 'None':
        circuit1_ppp_name = 'dummy@bband1.com'
        circuit1_ppp_pwd = 'dummy'
    
    if circuit1_ppp_name == 'None':
        circuit1_ppp_name = 'notrequired'
    
    # get provision port status
    provision_port_disable = str(tracker_sheet_obj.cell(row=tracker_row, column=provision_port_disable_col).value)
    if provision_port_disable == 'None':
        provision_port_disable = 'FALSE'
    else:
        provision_port_disable = 'TRUE'

    # get vlan 2 network
    vlan2_ipv4 = str(tracker_sheet_obj.cell(row=tracker_row, column=vlan2_col).value)
    if vlan2_ipv4 and '/' not in vlan2_ipv4:
        vlan2_ipv4 = vlan2_ipv4 + '/28'
    vlan2_ipv4 = ipaddress.ip_network(vlan2_ipv4, strict=False)

    # generate cctv nat from store number
    a = int(store_num) * 4

    octet3 = math.floor(int(a) / 256)
    octet4 = a - (octet3 * 256)

    cctv_nat = ipaddress.ip_address(f'172.19.{octet3}.{octet4}')

    # generate store networks from store number
    store_net_oct2, store_net_oct3, store_net_oct2_vlan70, store_net_oct2_vlan31, store_net_oct3_vlan31, store_net_oct2_vlan101 = store_nets(store_num)
    
    vlan60_ipv4 = str(tracker_sheet_obj.cell(row=tracker_row, column=vlan60_col).value)

    if store_type == 3 or store_type == 4:

        vlan60_ipv4 = ipaddress.ip_network(f'151.{store_net_oct2}.{store_net_oct3}.0/24')
        vlan20_ipv4 = ipaddress.ip_network(f'10.1{store_net_oct2}.{store_net_oct3}.128/25')
    else:
        if vlan60_ipv4 and '/' not in vlan60_ipv4:
            vlan60_ipv4 = vlan60_ipv4 + '/24'
        vlan60_ipv4 = ipaddress.ip_network(vlan60_ipv4, strict=False)
        vlan20_ipv4 = ipaddress.ip_network(f'{vlan60_ipv4.network_address.packed[0]}.1{vlan60_ipv4.network_address.packed[1]}.{vlan60_ipv4.network_address.packed[2]}.{vlan60_ipv4.network_address.packed[3]}/24')
    
    vlan70_ipv4 = ipaddress.ip_network(f'10.{store_net_oct2_vlan70}.{store_net_oct3}.0/24')
    vlan80_ipv4 = ipaddress.ip_network(f'192.168.100.0/24')
    vlan10_ipv4 = ipaddress.ip_network(f'10.1{store_net_oct2}.{store_net_oct3}.0/25')
    vlan30_ipv4 = ipaddress.ip_network(f'192.168.101.0/24')
    vlan31_ipv4 = ipaddress.ip_network(f'10.{store_net_oct2_vlan31}.{store_net_oct3_vlan31}.0/28')
    vlan40_ipv4 = ipaddress.ip_network(f'192.168.102.0/24')
    vlan100_ipv4 = ipaddress.ip_network(f'192.168.103.0/24')
    vlan101_ipv4 = ipaddress.ip_network(f'10.{store_net_oct2_vlan101}.{store_net_oct3}.224/27')
    vlan120_ipv4 = ipaddress.ip_network(f'192.168.104.0/24')

    # add the globally significant subnets (vrf 100, 700) to a python set to check for duplicates
    # checks for Vlan 42 - Wesley Media (not required as not globally significant - vrf 400)
    # checks for Vlan 192 - Cremators (not required as not globally significant - vrf 400)

    if store_type == 3 or store_type == 4:
        store_subnets = [str(vlan10_ipv4), str(vlan20_ipv4), str(vlan31_ipv4), str(vlan60_ipv4), str(vlan70_ipv4), str(vlan2_ipv4)]
    
    if store_type == 5 or store_type == 6:
        store_subnets = [str(vlan20_ipv4), str(vlan31_ipv4), str(vlan60_ipv4), str(vlan101_ipv4), str(vlan2_ipv4)]
    
    for subnet in store_subnets:
        if subnet in unique_subnets:
            print(f'Error: Duplicate subnet {subnet} found for store {store_num} row {tracker_row}  ... ABORTED - Please correct and re-run')
            tracker_row = tracker_row + 1
            sys.exit()
        else:
            unique_subnets.add(subnet)

    # print store networks for debugging

    print_nets = False
    if print_nets:
        print(f'Store {store_num} VLAN networks:')
        print(f'VLAN2: {vlan2_ipv4}')
        print(f'VLAN10: {vlan10_ipv4}')
        print(f'VLAN20: {vlan20_ipv4}')
        print(f'VLAN30: {vlan30_ipv4}')
        print(f'VLAN31: {vlan31_ipv4}')
        print(f'VLAN40: {vlan40_ipv4}')
        print(f'VLAN60: {vlan60_ipv4}')
        print(f'VLAN70: {vlan70_ipv4}')
        print(f'VLAN80: {vlan80_ipv4}')
        print(f'VLAN100: {vlan100_ipv4}')
        print(f'VLAN101: {vlan101_ipv4}')
        print(f'VLAN120: {vlan120_ipv4}')
        print('')


    # build the dictionary rows for router 1
    vmanage_dict['Device ID'].append("C1121X-8P-" + router1_serial)
    vmanage_dict['System IP'].append(str(router1_systemip))
    vmanage_dict['Host Name'].append(router1_hostname)
    vmanage_dict['Site Id'].append(site_id)
    vmanage_dict['Dual Stack IPv6 Default'].append('FALSE')
    vmanage_dict['Rollback Timer (sec)'].append('300')
    vmanage_dict['provision_port_disable'].append(provision_port_disable)
    vmanage_dict['vlan31_vrrp_pri'].append('110')
    vmanage_dict['vlan31_vrrp_ipv4'].append(str(vlan31_ipv4.network_address + 14))
    vmanage_dict['vlan31_ipv4'].append(str(vlan31_ipv4.network_address + 12))
    vmanage_dict['vlan31_mask'].append(str(vlan31_ipv4.netmask))
    vmanage_dict['vlan31_dhcp_net'].append(str(vlan31_ipv4.network_address))
    vmanage_dict['vlan31_dhcp_mask'].append(str(vlan31_ipv4.netmask))
    vmanage_dict['vlan31_dhcp_exclude'].append(f'{str(vlan31_ipv4.network_address + 7)}-{str(vlan31_ipv4.network_address + 14)}')
    vmanage_dict['vlan31_dhcp_gateway'].append(str(vlan31_ipv4.network_address + 14))
    vmanage_dict['vlan120_vrrp_pri'].append('110')
    vmanage_dict['vlan120_vrrp_ipv4'].append(str(vlan120_ipv4.network_address + 254))
    vmanage_dict['vlan120_ipv4'].append(str(vlan120_ipv4.network_address + 252))
    vmanage_dict['vlan120_mask'].append(str(vlan120_ipv4.netmask))
    vmanage_dict['vlan120_dhcp_exclude'].append(f'{str(vlan120_ipv4.network_address + 128)}-{str(vlan120_ipv4.network_address + 254)}')
    vmanage_dict['vlan100_vrrp_pri'].append('110')
    vmanage_dict['vlan100_vrrp_ipv4'].append(str(vlan100_ipv4.network_address + 254))
    vmanage_dict['vlan100_ipv4'].append(str(vlan100_ipv4.network_address + 252))
    vmanage_dict['vlan100_mask'].append(str(vlan100_ipv4.netmask))
    vmanage_dict['vlan100_dhcp_exclude'].append(f'{str(vlan100_ipv4.network_address + 128)}-{str(vlan100_ipv4.network_address + 254)}')
    vmanage_dict['vlan101_vrrp_pri'].append('110')
    vmanage_dict['vlan101_vrrp_ipv4'].append(str(vlan101_ipv4.network_address + 30))
    vmanage_dict['vlan101_ipv4'].append(str(vlan101_ipv4.network_address + 28))
    vmanage_dict['vlan101_mask'].append(str(vlan101_ipv4.netmask))
    vmanage_dict['vlan101_dhcp_net'].append(str(vlan101_ipv4.network_address))
    vmanage_dict['vlan101_dhcp_mask'].append(str(vlan101_ipv4.netmask))
    vmanage_dict['vlan101_dhcp_gateway'].append(str(vlan101_ipv4.network_address + 30))
    vmanage_dict['vlan101_dhcp_exclude'].append(f'{str(vlan101_ipv4.network_address + 14)}-{str(vlan101_ipv4.network_address + 30)}')
    vmanage_dict['vlan40_vrrp_pri'].append('110')
    vmanage_dict['vlan40_vrrp_ipv4'].append(str(vlan40_ipv4.network_address + 254))
    vmanage_dict['vlan40_ipv4'].append(str(vlan40_ipv4.network_address + 252))
    vmanage_dict['vlan40_mask'].append(str(vlan40_ipv4.netmask))
    vmanage_dict['vlan40_dhcp_exclude'].append(f'{str(vlan40_ipv4.network_address + 128)}-{str(vlan40_ipv4.network_address + 254)}')
    vmanage_dict['vlan30_vrrp_pri'].append('110')
    vmanage_dict['vlan30_vrrp_ipv4'].append(str(vlan30_ipv4.network_address + 254))
    vmanage_dict['vlan30_ipv4'].append(str(vlan30_ipv4.network_address + 252))
    vmanage_dict['vlan30_mask'].append(str(vlan30_ipv4.netmask))
    vmanage_dict['vlan30_dhcp_exclude'].append(f'{str(vlan30_ipv4.network_address + 128)}-{str(vlan30_ipv4.network_address + 254)}')
    vmanage_dict['vlan20_vrrp_pri'].append('110')
    vmanage_dict['vlan20_vrrp_ipv4'].append(str(vlan20_ipv4.network_address + 126))
    vmanage_dict['vlan20_ipv4'].append(str(vlan20_ipv4.network_address + 124))
    vmanage_dict['vlan20_mask'].append(str(vlan20_ipv4.netmask))
    vmanage_dict['vlan20_dhcp_net'].append(str(vlan20_ipv4.network_address))
    vmanage_dict['vlan20_dhcp_mask'].append(str(vlan20_ipv4.netmask))
    vmanage_dict['vlan20_dhcp_exclude'].append(f'{str(vlan20_ipv4.network_address + 64)}-{str(vlan20_ipv4.network_address + 126)}')
    vmanage_dict['vlan20_dhcp_gateway'].append(str(vlan20_ipv4.network_address + 126))
    vmanage_dict['vlan10_vrrp_pri'].append('110')
    vmanage_dict['vlan10_vrrp_ipv4'].append(str(vlan10_ipv4.network_address + 126))
    vmanage_dict['vlan10_ipv4'].append(str(vlan10_ipv4.network_address + 124))
    vmanage_dict['vlan10_mask'].append(str(vlan10_ipv4.netmask))
    vmanage_dict['vlan10_dhcp_net'].append(str(vlan10_ipv4.network_address))
    vmanage_dict['vlan10_dhcp_mask'].append(str(vlan10_ipv4.netmask))
    vmanage_dict['vlan10_dhcp_exclude'].append(f'{str(vlan10_ipv4.network_address + 64)}-{str(vlan10_ipv4.network_address + 126)}')
    vmanage_dict['vlan10_dhcp_gateway'].append(str(vlan10_ipv4.network_address + 126))
    vmanage_dict['vlan60_vrrp_pri'].append('110')
    vmanage_dict['vlan60_vrrp_ipv4'].append(str(vlan60_ipv4.network_address + 254))
    vmanage_dict['vlan60_ipv4'].append(str(vlan60_ipv4.network_address + 252))
    vmanage_dict['vlan60_mask'].append(str(vlan60_ipv4.netmask))
    vmanage_dict['vlan70_vrrp_pri'].append('110')
    vmanage_dict['vlan70_vrrp_ipv4'].append(str(vlan70_ipv4.network_address + 254))
    vmanage_dict['vlan70_ipv4'].append(str(vlan70_ipv4.network_address + 252))
    vmanage_dict['vlan70_mask'].append(str(vlan70_ipv4.netmask))
    vmanage_dict['vlan80_vrrp_pri'].append('110')
    vmanage_dict['vlan80_vrrp_ipv4'].append(str(vlan80_ipv4.network_address + 1))
    vmanage_dict['vlan80_ipv4'].append(str(vlan80_ipv4.network_address + 252))
    vmanage_dict['vlan80_mask'].append(str(vlan80_ipv4.netmask))
    vmanage_dict['tloc_next_hop'].append(str('192.168.12.2'))
    vmanage_dict['tloc_bandwidth_up'].append(str(int(circuit1_bw_up * 1000)))
    vmanage_dict['tloc_bandwidth_down'].append(str(int(circuit1_bw_down * 1000)))
    vmanage_dict['wan_bandwidth_up'].append(str(int(circuit1_bw_up * 1000)))
    vmanage_dict['wan_bandwidth_down'].append(str(int(circuit1_bw_down * 1000)))
    vmanage_dict['wan_desc'].append(f'{circuit1_ref} - {circuit1_type} via {circuit1_provider}')
    vmanage_dict['ethpppoe_chapHost'].append(circuit1_ppp_name)
    vmanage_dict['ethpppoe_chapPwd'].append(circuit1_ppp_pwd)
    vmanage_dict['wan_color'].append(router1_wan_color)
    vmanage_dict['ethpppoe_ipsecPrefer'].append('111')
    vmanage_dict['wan_shapingRate'].append(int(circuit1_bw_up * 1000))
    vmanage_dict['wan_track_addr'].append('1.1.1.1')
    vmanage_dict['wan_track_addr_tloc'].append('208.67.222.222')
    vmanage_dict['loopback0_ipv4'].append(str(router1_systemip))
    vmanage_dict['loopback0_mask'].append('255.255.255.255')
    vmanage_dict['lan_vpn_100_nat_1_rangeStart'].append(str(cctv_nat))
    vmanage_dict['lan_vpn_100_nat_1_rangeEnd'].append(str(cctv_nat + 1))
    vmanage_dict['lan_vpn_100_staticNat_1_translatedSourceIp'].append(str(cctv_nat))
    vmanage_dict['lan_vpn_100_staticNat_2_translatedSourceIp'].append(str(cctv_nat + 1))
    vmanage_dict['vlan2_vrrp_pri'].append('110')
    vmanage_dict['vlan2_vrrp_ipv4'].append(str(vlan2_ipv4.network_address + 14))
    vmanage_dict['vlan2_ipv4'].append(str(vlan2_ipv4.network_address + 12))
    vmanage_dict['vlan2_mask'].append(str(vlan2_ipv4.netmask))
    vmanage_dict['vlan2_dhcp_net'].append(str(vlan2_ipv4.network_address))
    vmanage_dict['vlan2_dhcp_mask'].append(str(vlan2_ipv4.netmask))
    vmanage_dict['vlan2_dhcp_exclude'].append(f'{str(vlan2_ipv4.network_address + 7)}-{str(vlan2_ipv4.network_address + 14)}')
    vmanage_dict['vlan2_dhcp_gateway'].append(str(vlan2_ipv4.network_address + 14))
    vmanage_dict['cloudSaaSDeviceRole_variable'].append('dia')
    vmanage_dict['cloudSaaSVpnType_variable'].append('service-vpn')
    vmanage_dict['cloudSaasTlocList_variable'].append('all')
    vmanage_dict['cloudSaasSigTunnelList_variable'].append('')
    vmanage_dict['cloudSaasSigEnabled_variable'].append('FALSE')
    vmanage_dict['cloudSaasInterfaceList_variable'].append('')
    vmanage_dict['cloudSaasLBEnabled_variable'].append('TRUE')
    vmanage_dict['cloudSaasLoss_variable'].append(5)
    vmanage_dict['cloudSaasLatency_variable'].append(100)
    vmanage_dict['cloudSaasSourceIpBased_variable'].append('TRUE')
    vmanage_dict['qos_Interface_1'].append(str(interface1))




    # if we have a router 2 build the dictionary rows for router 2
    if router2_serial != 'NONE':
        vmanage_dict['Device ID'].append("C1121X-8P-" + router2_serial)
        vmanage_dict['System IP'].append(str(router2_systemip))
        vmanage_dict['Host Name'].append(router2_hostname)
        vmanage_dict['Site Id'].append(site_id)
        vmanage_dict['Dual Stack IPv6 Default'].append('FALSE')
        vmanage_dict['Rollback Timer (sec)'].append('300')
        vmanage_dict['provision_port_disable'].append(provision_port_disable)
        vmanage_dict['vlan31_vrrp_pri'].append('100')
        vmanage_dict['vlan31_vrrp_ipv4'].append(str(vlan31_ipv4.network_address + 14))
        vmanage_dict['vlan31_ipv4'].append(str(vlan31_ipv4.network_address + 13))
        vmanage_dict['vlan31_mask'].append(str(vlan31_ipv4.netmask))
        vmanage_dict['vlan31_dhcp_net'].append(str(vlan31_ipv4.network_address))
        vmanage_dict['vlan31_dhcp_mask'].append(str(vlan31_ipv4.netmask))
        vmanage_dict['vlan31_dhcp_exclude'].append(f'{str(vlan31_ipv4.network_address + 1)}-{str(vlan31_ipv4.network_address + 6)}";"{str(vlan31_ipv4.network_address + 12)}-{str(vlan31_ipv4.network_address + 14)}')
        vmanage_dict['vlan31_dhcp_gateway'].append(str(vlan31_ipv4.network_address + 14))
        vmanage_dict['vlan120_vrrp_pri'].append('100')
        vmanage_dict['vlan120_vrrp_ipv4'].append(str(vlan120_ipv4.network_address + 254))
        vmanage_dict['vlan120_ipv4'].append(str(vlan120_ipv4.network_address + 253))
        vmanage_dict['vlan120_mask'].append(str(vlan120_ipv4.netmask))
        vmanage_dict['vlan120_dhcp_exclude'].append(f'{str(vlan120_ipv4.network_address + 1)}-{str(vlan120_ipv4.network_address + 127)}";"{str(vlan120_ipv4.network_address + 252)}-{str(vlan120_ipv4.network_address + 254)}')
        vmanage_dict['vlan100_vrrp_pri'].append('100')
        vmanage_dict['vlan100_vrrp_ipv4'].append(str(vlan100_ipv4.network_address + 254))
        vmanage_dict['vlan100_ipv4'].append(str(vlan100_ipv4.network_address + 253))
        vmanage_dict['vlan100_mask'].append(str(vlan100_ipv4.netmask))
        vmanage_dict['vlan100_dhcp_exclude'].append(f'{str(vlan100_ipv4.network_address + 1)}-{str(vlan100_ipv4.network_address + 127)}";"{str(vlan100_ipv4.network_address + 252)}-{str(vlan100_ipv4.network_address + 254)}')
        vmanage_dict['vlan101_vrrp_pri'].append('100')
        vmanage_dict['vlan101_vrrp_ipv4'].append(str(vlan101_ipv4.network_address + 30))
        vmanage_dict['vlan101_ipv4'].append(str(vlan101_ipv4.network_address + 29))
        vmanage_dict['vlan101_mask'].append(str(vlan101_ipv4.netmask))
        vmanage_dict['vlan101_dhcp_net'].append(str(vlan101_ipv4.network_address))
        vmanage_dict['vlan101_dhcp_mask'].append(str(vlan101_ipv4.netmask))
        vmanage_dict['vlan101_dhcp_gateway'].append(str(vlan101_ipv4.network_address + 30))
        vmanage_dict['vlan101_dhcp_exclude'].append(f'{str(vlan101_ipv4.network_address + 1)}-{str(vlan101_ipv4.network_address + 13)}";"{str(vlan101_ipv4.network_address + 28)}-{str(vlan101_ipv4.network_address + 30)}')
        vmanage_dict['vlan40_vrrp_pri'].append('100')
        vmanage_dict['vlan40_vrrp_ipv4'].append(str(vlan40_ipv4.network_address + 254))
        vmanage_dict['vlan40_ipv4'].append(str(vlan40_ipv4.network_address + 253))
        vmanage_dict['vlan40_mask'].append(str(vlan40_ipv4.netmask))
        vmanage_dict['vlan40_dhcp_exclude'].append(f'{str(vlan40_ipv4.network_address + 1)}-{str(vlan40_ipv4.network_address + 127)}";"{str(vlan40_ipv4.network_address + 252)}-{str(vlan40_ipv4.network_address + 254)}')
        vmanage_dict['vlan30_vrrp_pri'].append('100')
        vmanage_dict['vlan30_vrrp_ipv4'].append(str(vlan30_ipv4.network_address + 254))
        vmanage_dict['vlan30_ipv4'].append(str(vlan30_ipv4.network_address + 253))
        vmanage_dict['vlan30_mask'].append(str(vlan30_ipv4.netmask))
        vmanage_dict['vlan30_dhcp_exclude'].append(f'{str(vlan30_ipv4.network_address + 1)}-{str(vlan30_ipv4.network_address + 127)}";"{str(vlan30_ipv4.network_address + 252)}-{str(vlan30_ipv4.network_address + 254)}')
        vmanage_dict['vlan20_vrrp_pri'].append('100')
        vmanage_dict['vlan20_vrrp_ipv4'].append(str(vlan20_ipv4.network_address + 126))
        vmanage_dict['vlan20_ipv4'].append(str(vlan20_ipv4.network_address + 125))
        vmanage_dict['vlan20_mask'].append(str(vlan20_ipv4.netmask))
        vmanage_dict['vlan20_dhcp_net'].append(str(vlan20_ipv4.network_address))
        vmanage_dict['vlan20_dhcp_mask'].append(str(vlan20_ipv4.netmask))
        vmanage_dict['vlan20_dhcp_exclude'].append(f'{str(vlan20_ipv4.network_address + 1)}-{str(vlan20_ipv4.network_address + 63)}";"{str(vlan20_ipv4.network_address + 124)}-{str(vlan20_ipv4.network_address + 126)}')
        vmanage_dict['vlan20_dhcp_gateway'].append(str(vlan20_ipv4.network_address + 126))
        vmanage_dict['vlan10_vrrp_pri'].append('100')
        vmanage_dict['vlan10_vrrp_ipv4'].append(str(vlan10_ipv4.network_address + 126))
        vmanage_dict['vlan10_ipv4'].append(str(vlan10_ipv4.network_address + 125))
        vmanage_dict['vlan10_mask'].append(str(vlan10_ipv4.netmask))
        vmanage_dict['vlan10_dhcp_net'].append(str(vlan10_ipv4.network_address))
        vmanage_dict['vlan10_dhcp_mask'].append(str(vlan10_ipv4.netmask))
        vmanage_dict['vlan10_dhcp_exclude'].append(f'{str(vlan10_ipv4.network_address + 1)}-{str(vlan10_ipv4.network_address + 63)}";"{str(vlan10_ipv4.network_address + 124)}-{str(vlan10_ipv4.network_address + 126)}')
        vmanage_dict['vlan10_dhcp_gateway'].append(str(vlan10_ipv4.network_address + 126))
        vmanage_dict['vlan60_vrrp_pri'].append('100')
        vmanage_dict['vlan60_vrrp_ipv4'].append(str(vlan60_ipv4.network_address + 254))
        vmanage_dict['vlan60_ipv4'].append(str(vlan60_ipv4.network_address + 253))
        vmanage_dict['vlan60_mask'].append(str(vlan60_ipv4.netmask))
        vmanage_dict['vlan70_vrrp_pri'].append('100')
        vmanage_dict['vlan70_vrrp_ipv4'].append(str(vlan70_ipv4.network_address + 254))
        vmanage_dict['vlan70_ipv4'].append(str(vlan70_ipv4.network_address + 253))
        vmanage_dict['vlan70_mask'].append(str(vlan70_ipv4.netmask))
        vmanage_dict['vlan80_vrrp_pri'].append('100')
        vmanage_dict['vlan80_vrrp_ipv4'].append(str(vlan80_ipv4.network_address + 1))
        vmanage_dict['vlan80_ipv4'].append(str(vlan80_ipv4.network_address + 253))
        vmanage_dict['vlan80_mask'].append(str(vlan80_ipv4.netmask))
        vmanage_dict['tloc_next_hop'].append(str('192.168.21.1'))
        vmanage_dict['tloc_bandwidth_up'].append(str(int(circuit2_bw_up * 1000)))
        vmanage_dict['tloc_bandwidth_down'].append(str(int(circuit2_bw_down * 1000)))
        vmanage_dict['wan_bandwidth_up'].append(str(int(circuit2_bw_up * 1000)))
        vmanage_dict['wan_bandwidth_down'].append(str(int(circuit2_bw_down * 1000)))
        vmanage_dict['wan_desc'].append(f'{circuit2_ref} - {circuit2_type} via {circuit2_provider}')
        vmanage_dict['ethpppoe_chapHost'].append(circuit2_ppp_name)
        vmanage_dict['ethpppoe_chapPwd'].append(circuit2_ppp_pwd)
        vmanage_dict['wan_color'].append(router2_wan_color)
        vmanage_dict['ethpppoe_ipsecPrefer'].append('0')
        vmanage_dict['wan_shapingRate'].append(int(circuit2_bw_up * 1000))
        vmanage_dict['wan_track_addr'].append('208.67.222.222')
        vmanage_dict['wan_track_addr_tloc'].append('1.1.1.1')
        vmanage_dict['loopback0_ipv4'].append(str(router2_systemip))
        vmanage_dict['loopback0_mask'].append('255.255.255.255')
        vmanage_dict['lan_vpn_100_nat_1_rangeStart'].append(str(cctv_nat))
        vmanage_dict['lan_vpn_100_nat_1_rangeEnd'].append(str(cctv_nat + 1))
        vmanage_dict['lan_vpn_100_staticNat_1_translatedSourceIp'].append(str(cctv_nat))
        vmanage_dict['lan_vpn_100_staticNat_2_translatedSourceIp'].append(str(cctv_nat + 1))
        vmanage_dict['vlan2_vrrp_pri'].append('100')
        vmanage_dict['vlan2_vrrp_ipv4'].append(str(vlan2_ipv4.network_address + 14))
        vmanage_dict['vlan2_ipv4'].append(str(vlan2_ipv4.network_address + 13))
        vmanage_dict['vlan2_mask'].append(str(vlan2_ipv4.netmask))
        vmanage_dict['vlan2_dhcp_net'].append(str(vlan2_ipv4.network_address))
        vmanage_dict['vlan2_dhcp_mask'].append(str(vlan2_ipv4.netmask))
        vmanage_dict['vlan2_dhcp_exclude'].append(f'{str(vlan2_ipv4.network_address + 1)}-{str(vlan2_ipv4.network_address + 6)}";"{str(vlan2_ipv4.network_address + 12)}-{str(vlan2_ipv4.network_address + 14)}')
        vmanage_dict['vlan2_dhcp_gateway'].append(str(vlan2_ipv4.network_address + 14))
        vmanage_dict['cloudSaaSDeviceRole_variable'].append('dia')
        vmanage_dict['cloudSaaSVpnType_variable'].append('service-vpn')
        vmanage_dict['cloudSaasTlocList_variable'].append('all')
        vmanage_dict['cloudSaasSigTunnelList_variable'].append('')
        vmanage_dict['cloudSaasSigEnabled_variable'].append('FALSE')
        vmanage_dict['cloudSaasInterfaceList_variable'].append('')
        vmanage_dict['cloudSaasLBEnabled_variable'].append('TRUE')
        vmanage_dict['cloudSaasLoss_variable'].append(5)
        vmanage_dict['cloudSaasLatency_variable'].append(100)
        vmanage_dict['cloudSaasSourceIpBased_variable'].append('TRUE')
        vmanage_dict['qos_Interface_1'].append(str(interface2))

    tracker_row = tracker_row + 1

# end of main loop

# perform postcode lookups to obtain GPS coords
print('\nPerforming postcode lookups ...\n')

# Break the postcode list into chunks of 100 as the API has a max 100 limit

latlist = []
longlist = []

while len(postcode_list) > 100:
    postcode_max100 = postcode_list[0:100]
    postcode_result = postcode_api(postcode_max100)
    postcode_df = pd.json_normalize(postcode_result.json()['result'],sep='_')
    # update the csv dictionary with the lat and long values returned by the API
    latlist = latlist + (postcode_df['result_latitude'].to_list())
    longlist = longlist+ (postcode_df['result_longitude'].to_list())
    # remove the postcodes we have lookedup from the list
    postcode_list = postcode_list[100:]

if len(postcode_list) > 0:
    postcode_result = postcode_api(postcode_list)
    postcode_df = pd.json_normalize(postcode_result.json()['result'],sep='_')
    latlist = latlist + (postcode_df['result_latitude'].to_list())
    longlist = longlist+ (postcode_df['result_longitude'].to_list())
    
# update the csv dictionary with the lat and long values returned by the API
vmanage_dict['basic_gpsl_latitude'] = latlist
vmanage_dict['basic_gpsl_longitude'] = longlist

# uncomment to print the dictionary for debugging
#print(json.dumps(vmanage_dict, indent=1))

# create the dataframe from the dictionary we built

# uncomment the following lines to pprint the dictionary - ValueError: All arrays must be of the same length means one or more of the dictonary lists is the wrong length compared to the rest
#pprint.pprint(vmanage_dict)
df = pd.DataFrame(vmanage_dict)

# write the dataframe to a csv ready for import into vManage
try:
    df.to_csv('/mnt/c/Users/nick.oneill/Downloads/vmanage-import-sc.csv', index=False)
except PermissionError:
    print('*' * 120,'\nError: vmanage-import-sc.csv is open in another application - please close and re-run the script\n','*' * 120)
    exit()

# all done
print('vmanage-import-sc.csv has been created :)\n')